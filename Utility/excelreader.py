import openpyxl


def read_excel():
	wb = openpyxl.load_workbook("./TestData.xlsx")
	ws = wb.active
	data = []

	for row in range(ws.max_row):
		username = ws['A' + str(row + 1)].value
		pwd = ws['B' + str(row + 1)].value
		if username is None:
			username = ""
		if pwd is None:
			pwd = ""
		data.append([username, pwd])

	return data




